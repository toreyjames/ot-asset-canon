#!/usr/bin/env python3

import json
import math
import os
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path

import requests
from openpyxl import load_workbook


DEFAULT_URL = "https://eta-publications.lbl.gov/sites/default/files/2025-08/lbnl_ix_queue_data_file_thru2024_v2.xlsx"
DEFAULT_SOURCE_PAGE = "https://emp.lbl.gov/queues"


def excel_date_to_iso(value):
    if value in (None, "", "NA"):
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc).isoformat()
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(value):
            return None
        base = datetime(1899, 12, 30, tzinfo=timezone.utc)
        return (base + timedelta(days=float(value))).isoformat()
    if isinstance(value, str):
        value = value.strip()
        if not value or value.upper() == "NA":
            return None
        try:
            return datetime.fromisoformat(value).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            return None
    return None


def as_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return None if not text or text.upper() == "NA" else text


def as_number(value):
    if value in (None, "", "NA"):
        return None
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    return parsed if math.isfinite(parsed) else None


def fetch_workbook(source):
    if source.startswith("http://") or source.startswith("https://"):
        response = requests.get(
            source,
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": DEFAULT_SOURCE_PAGE,
            },
            timeout=120,
        )
        response.raise_for_status()
        target = Path("/tmp/interconnection-queue.xlsx")
        target.write_bytes(response.content)
        return target
    return Path(source)


def main():
    source = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_URL
    row_limit = int(sys.argv[2]) if len(sys.argv) > 2 else 250
    min_mw = float(sys.argv[3]) if len(sys.argv) > 3 else 100.0

    workbook_path = fetch_workbook(source)
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["03. Complete Queue Data"]

    rows = ws.iter_rows(min_row=2, values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]

    accepted_statuses = {"active", "operational", "suspended"}
    parsed = []

    for row in rows:
      record = dict(zip(headers, row))
      status = (as_text(record.get("q_status")) or "").lower()
      if status not in accepted_statuses:
          continue

      mw_total = sum(
          value or 0
          for value in (
              as_number(record.get("mw1")),
              as_number(record.get("mw2")),
              as_number(record.get("mw3")),
          )
      )
      if mw_total < min_mw:
          continue

      q_id = as_text(record.get("q_id"))
      if not q_id:
          continue

      parsed.append(
          {
              "id": q_id,
              "qStatus": status,
              "qDate": excel_date_to_iso(record.get("q_date")),
              "proposedDate": excel_date_to_iso(record.get("prop_date")),
              "operationalDate": excel_date_to_iso(record.get("on_date")),
              "withdrawnDate": excel_date_to_iso(record.get("wd_date")),
              "agreementDate": excel_date_to_iso(record.get("ia_date")),
              "countyName": as_text(record.get("county")),
              "state": as_text(record.get("state")),
              "countyFips": (
                  str(int(record.get("fips_codes"))).zfill(5)
                  if as_number(record.get("fips_codes")) is not None
                  and len(str(int(record.get("fips_codes"))).zfill(5)) == 5
                  else None
              ),
              "poiName": as_text(record.get("poi_name")),
              "region": as_text(record.get("region")),
              "projectName": as_text(record.get("project_name")),
              "utility": as_text(record.get("utility")),
              "entity": as_text(record.get("entity")),
              "developer": as_text(record.get("developer")),
              "service": as_text(record.get("service")),
              "projectCategory": as_text(record.get("project_type")),
              "typeClean": as_text(record.get("type_clean")),
              "type1": as_text(record.get("type1")),
              "type2": as_text(record.get("type2")),
              "type3": as_text(record.get("type3")),
              "capacityMw": mw_total,
              "queueYear": int(record.get("q_year")) if as_number(record.get("q_year")) is not None else None,
              "proposedYear": int(record.get("prop_year")) if as_number(record.get("prop_year")) is not None else None,
              "sourceUrl": source if source.startswith("http") else DEFAULT_SOURCE_PAGE,
          }
      )

    status_rank = {"active": 3, "suspended": 2, "operational": 1}
    parsed.sort(
        key=lambda row: (
            status_rank.get(row.get("qStatus"), 0),
            row.get("proposedYear") or 0,
            row.get("queueYear") or 0,
            row.get("capacityMw") or 0,
        ),
        reverse=True,
    )
    if row_limit:
        parsed = parsed[:row_limit]

    print(json.dumps({"records": parsed}))


if __name__ == "__main__":
    main()
