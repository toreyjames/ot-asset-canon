#!/usr/bin/env python3

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) < 2:
      raise SystemExit("Usage: parse_cbsa_delineation.py <workbook_path>")

    workbook_path = Path(sys.argv[1])
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    rows = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        cbsa_code = row[0]
        cbsa_title = row[3]
        area_type = row[4]
        county_name = row[7]
        state_name = row[8]
        state_fips = row[9]
        county_fips = row[10]
        county_role = row[11]

        if not cbsa_code or not state_fips or not county_fips:
            continue

        rows.append(
            {
                "countyFips": f"{str(state_fips).zfill(2)}{str(county_fips).zfill(3)}",
                "stateFips": str(state_fips).zfill(2),
                "cbsaCode": str(cbsa_code).zfill(5),
                "cbsaName": str(cbsa_title).strip() if cbsa_title else None,
                "countyName": str(county_name).strip() if county_name else None,
                "stateName": str(state_name).strip() if state_name else None,
                "areaType": str(area_type).strip() if area_type else None,
                "countyRole": str(county_role).strip() if county_role else None,
            }
        )

    print(json.dumps(rows))


if __name__ == "__main__":
    main()
