#!/usr/bin/env python3

import json
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: parse_eia860_plant.py <zip_path>")

    zip_path = Path(sys.argv[1])
    with zipfile.ZipFile(zip_path) as archive:
        target_name = None
        for name in archive.namelist():
            if name.lower().endswith("2___plant_y2024.xlsx"):
                target_name = name
                break
        if not target_name:
            raise SystemExit("Could not find 2___Plant_Y2024.xlsx in archive")

        extract_path = Path("/tmp/eia860_plant_y2024.xlsx")
        extract_path.write_bytes(archive.read(target_name))

    workbook = load_workbook(extract_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header = [str(cell).strip() if cell is not None else "" for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))]
    rows = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
      item = {}
      for index, key in enumerate(header):
          if not key:
              continue
          value = row[index] if index < len(row) else None
          item[key] = value
      plant_code = item.get("Plant Code")
      state_code = item.get("State")
      county_name = item.get("County")
      if not plant_code or not state_code:
          continue
      rows.append(
          {
              "plantCode": str(plant_code),
              "plantName": item.get("Plant Name"),
              "streetAddress": item.get("Street Address"),
              "city": item.get("City"),
              "state": str(state_code).strip(),
              "zip": str(item.get("Zip")).split(".")[0] if item.get("Zip") is not None else None,
              "countyName": str(county_name).strip() if county_name else None,
              "latitude": item.get("Latitude"),
              "longitude": item.get("Longitude"),
              "utilityName": item.get("Utility Name"),
              "utilityId": str(item.get("Utility ID")) if item.get("Utility ID") is not None else None,
              "sectorName": item.get("Sector Name"),
              "naicsCode": str(item.get("Primary Purpose (NAICS Code)")) if item.get("Primary Purpose (NAICS Code)") is not None else None,
            }
      )

    print(json.dumps(rows))


if __name__ == "__main__":
    main()
